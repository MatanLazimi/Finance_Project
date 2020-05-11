import openpyxl
from datetime import date
#Matan Lazimi - 203987607
#Elhanan Ben Shabu - 302642889
#Polina Ovars - 336382114
def calculateAge(birthDate):
    today = date.today()
    age = today.year - birthDate.year - ((today.month, today.day) < (birthDate.month, birthDate.day))
    return age

def death_calc(gender,age):
    path = "death.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    if (gender == 0):
        sheet_obj = wb_obj['male']
        for k in range(3, 96):
            cell_obj0 = sheet_obj.cell(row=k, column=2)
            cell_obj1 = sheet_obj.cell(row=k, column=6)
            if age == cell_obj0.value:
                return cell_obj1.value
    else:
        sheet_obj = wb_obj['female']
        for k in range(3, 96):
            cell_obj2 = sheet_obj.cell(row=k, column=2)
            cell_obj3 = sheet_obj.cell(row=k, column=6)
            if age == cell_obj2.value:
                return cell_obj3.value
######################################################################################################
def calc_discountrate(years_of_work):
    path = "data8.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj['dis']
    for k in range(5, 54):
        cell_obj1 = sheet_obj.cell(row=k, column=2)
        cell_obj2 = sheet_obj.cell(row=k, column=3)
        if years_of_work == cell_obj1.value:
            return cell_obj2.value


def calculation(seniority, non_article14, salary_growth_rate, last_salary, retirement_years, resignation, dismissal, asset, asset_flag, death_precentage, not_left, name, id, article14_precentage, age):
    calc = 0
    #check if there is article 14
    if seniority == non_article14:
        #if not
        for t in range(0, retirement_years):
            ################################################################################
            mone_dissmisal = ((1+salary_growth_rate)**(t+0.5)) * (not_left ** t) * dismissal
            mechane_dissmisal = ((1+calc_discountrate(t+1))**(t+0.5))
            calc_dissmisal = last_salary * seniority * (mone_dissmisal/mechane_dissmisal)
            ################################################################################
            if asset_flag:
                calc_res_asset = asset * resignation
            else:
                calc_res_asset = last_salary * resignation
            ################################################################################
            mone_death = ((1+salary_growth_rate)**(t+0.5)) * (not_left ** t) * death_precentage
            mechane_death = ((1+calc_discountrate(t+1))**(t+0.5))
            calc_death = last_salary * seniority * (mone_death/mechane_death)
            ################################################################################
            calc = calc + calc_dissmisal + calc_res_asset + calc_death
        #print("NAME === ",name ,"ID === ",id," ANSWER === ",calc)
    else:
        #if yes article 14
        if article14_precentage != 1:
            new_seniority = (1 - article14_precentage) * seniority
        else:
            new_seniority = seniority
        for t in range(0, retirement_years):
            ################################################################################
            mone_dissmisal = ((1+salary_growth_rate)**(t+0.5)) * (not_left ** t) * dismissal
            mechane_dissmisal = ((1+calc_discountrate(t+1))**(t+0.5))
            calc_dissmisal = last_salary * new_seniority * (mone_dissmisal/mechane_dissmisal)
            ################################################################################
            if non_article14 >= 1:
                if asset_flag:
                    calc_res_asset = asset * ( not_left**t ) * resignation
                else:
                    calc_res_asset = last_salary * ( not_left**t ) * resignation
            else:
                calc_res_asset = 0
            ################################################################################
            mone_death = ((1+salary_growth_rate)**(t+0.5)) * (not_left ** t) * death_precentage
            mechane_death = ((1+calc_discountrate(t+1))**(t+0.5))
            calc_death = last_salary * new_seniority * (mone_death/mechane_death)
            ################################################################################
            calc = calc + calc_dissmisal + calc_res_asset + calc_death
    print("ID === {1}\tNAME === {0}\tANSWER === {2}\tAGE === {3} \n".format(name,id,calc,age),end="")
    file1 = open("output.txt", "a", encoding='utf-8')
    # \n is placed to indicate EOL (End of Line)
    str = "ID === {1}\tNAME === {0}\tANSWER === {2}\tAGE === {3} \n".format(name,id,calc,age)
    file1.write(str)
    file1.close()  # to change file access modes
    return calc
##############################################################################################
def main():
    sum = 0
    path = "data8.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
#---------------Salaryyyyyyyyyyyyyyyyyy--------------------------------#
#    for k in range(3, 152):
#        cell_obj = sheet_obj.cell(row=k, column=7)
#        print(float(cell_obj.value))
######################################################################


#---------------CalcSenyority--------------------------------#
    for k in range(3,152):
        cell_obj0 = sheet_obj.cell(row=k, column=2)
        cell_obj1 = sheet_obj.cell(row=k, column=6)
        cell_obj2 = sheet_obj.cell(row=k, column=12)
        cell_obj3 = sheet_obj.cell(row=k, column=4)
        cell_obj4 = sheet_obj.cell(row=k, column=1)
        cell_obj5 = sheet_obj.cell(row=k, column=5)
        cell_obj6 = sheet_obj.cell(row=k, column=7)
        cell_obj7 = sheet_obj.cell(row=k, column=8)
        cell_obj8 = sheet_obj.cell(row=k, column=10)
        cell_obj9 = sheet_obj.cell(row=k, column=9)
        if cell_obj2.value is None or cell_obj2.value == "-" :
            seniority = (date.today() - (cell_obj1.value).date()).days/365.25
            if cell_obj7.value is None or cell_obj7.value == "-":
                non_article14 = seniority
            else:
                non_article14 = ((cell_obj7.value).date() - (cell_obj1.value).date()).days/365.25
            """Time for years of article 14"""
            article14 = seniority - non_article14
            """man is 0, female is 1."""
            if cell_obj3.value == "M":
                gender = 0
            else:
                gender = 1
            id = int(cell_obj4.value)
            if id % 2 == 0:
                salary_growth_rate = 0.04
            else:
                salary_growth_rate = 0.02

            age = calculateAge(cell_obj5.value.date())
            last_salary = float(cell_obj6.value)
            name = cell_obj0.value
            #########################################
            if gender == 1:
                retirement_years = 62 - age
            else:
                retirement_years = 67 - age
            #########################################
            if 18 <= age and 29 >= age:
                resignation = 0.20
                dismissal = 0.07
            elif 30 <= age and 39 >= age:
                resignation = 0.13
                dismissal = 0.05
            elif 40 <= age and 49 >= age:
                resignation = 0.10
                dismissal = 0.04
            elif 50 <= age and 59 >= age:
                resignation = 0.07
                dismissal = 0.03
            elif 60 <= age and 67 >= age:
                resignation = 0.03
                dismissal = 0.02
            #########################################
            #print("resi - ",resignation," diss - ",dismissal,age)
            asset = float(cell_obj8.value)
            #print("name - ", name, " asset - ",asset," age - ",age)
            ################elhanan################################
            if asset == 0.0:
                asset_flag = False
            else:
                asset_flag = True
            #######################################################
            death_precentage = death_calc(gender, age)
            not_left = 1 - (resignation + dismissal + death_precentage)
            if cell_obj9.value is not None:
                article14_precentage = int(cell_obj9.value)/100
            else:
                article14_precentage = 0
            #print("NUM == ", not_left)
            #print("name - ",name," ID - ",id," gender - ",gender," age - ",age," salary - ",last_salary," seniority - ",seniority," non_article14 - ",non_article14," article14 - ",article14," rate - ",salary_growth_rate)

            calc = calculation(seniority, non_article14, salary_growth_rate, last_salary, retirement_years, resignation, dismissal, asset, asset_flag, death_precentage, not_left, name, id, article14_precentage, age)
            sum = sum + calc
        else:
            pass
            """
            print('left: ',(cell_obj2.value.date() - cell_obj1.value.date()).days/365.25,end=" ")
            """
    # \n is placed to indicate EOL (End of Line)
    print(" SUM ALL == ", sum)
    file1 = open("output.txt", "a", encoding='utf-8')
    file1.write("SUM ALL == {0}".format(sum))
    file1.close()  # to change file access modes
######################################################################


main()